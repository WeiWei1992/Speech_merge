import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import os
import time
from openpyxl.styles import PatternFill
from openpyxl.styles import Color, Font, Alignment

import logging
import logging.config
CON_LOG='log.conf'
logging.config.fileConfig(CON_LOG)
logging=logging.getLogger()

def handle_excel(filepath=None):
    if not filepath:
        filepath='语料.xlsx'

    wb=load_workbook(filepath)
    sheet = wb.active
    i = sheet.max_row
    j=sheet.max_column
    # print("i： ",i)
    # print("j:  ",j)
    yuliao_column=0
    result_column=0

    for jj in range(1,j):
        if sheet.cell(1,jj).value=="语料":
            yuliao_column=jj
        elif sheet.cell(1,jj).value=="结果":
            result_column=jj
    result=[]
    #result_tmp=[]
    for ii in range(2,i+1):
        result_tmp = []
        yuliao_value=sheet.cell(ii,yuliao_column).value
        # print("ii: ",ii)
        # print("yuliao_column:",yuliao_column)
        # print(yuliao_value)
        result_tmp.append(ii)
        result_tmp.append(yuliao_column)
        result_tmp.append(yuliao_value)

        result.append(result_tmp)

    logging.info("结果以字典格式保存：" +str(result))
    return result

    #print(result)

def write_excel(action,result,filepath=None):
    if not filepath:
        filepath="语料.xlsx"
    wb=load_workbook(filepath)
    sheet=wb.active
    red_fill = PatternFill(fill_type='solid', fgColor="ff441f")
    green_file = PatternFill(fill_type='solid', fgColor="16feb5")
    if result:
        sheet.cell(row=action[0],column=action[1]+1,value=result).fill=green_file
    else:
        sheet.cell(row=action[0], column=action[1] + 1, value=result).fill = red_fill

    #sheet.cell(action[0],action[1]+1).value=result
    wb.save(filepath)



if __name__=="__main__":
    handle_excel()
    action=[2, 1, '今天天气怎么样']
    result=False
    write_excel(action,result)
